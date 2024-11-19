import cv2
import numpy as np
import loguru
import csv
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill



def ascii_to_image(input_csv_path, output_image_path):
    """
    從 CSV 格式的 ASCII 字符畫轉換為灰度圖片並保存為 PNG 格式。

    參數:
    - input_csv_path (str): CSV 文件的路徑。
    - output_image_path (str): 要保存的圖片檔案路徑。
    """
    # ASCII 字符與亮度對應表
    #N_str = '@%$&^!+=-. '
    N_str = '@%$&#*^!+  ' #New
    char_to_brightness = {char: idx * (255 / (len(N_str) - 1)) for idx, char in enumerate(N_str)}

    # 讀取 ASCII 字符畫 CSV
    with open(input_csv_path, 'r') as file:
        reader = csv.reader(file)
        ascii_art = [row for row in reader]

    # 確定圖像大小
    re_height = len(ascii_art)
    re_width = max(len(row) for row in ascii_art)
    gray_img = np.zeros((re_height, re_width), dtype=np.uint8)

    # 將每個字符轉換為對應的灰度值
    for j, row in enumerate(ascii_art):
        for i, char in enumerate(row):
            gray_img[j, i] = char_to_brightness.get(char, 0)
            loguru.logger.debug(f'process_to_graphic{j, i, char, gray_img}')

    # 儲存圖片
    cv2.imwrite(output_image_path, gray_img)
    loguru.logger.info(f"圖片已保存至 {output_image_path}")

def grayscale(rgb):
        return np.dot(rgb[...,:3], [0.299, 0.587, 0.114])

def image_to_asciitxt(file_root):
    file_name = file_root  ### Input your Image
    img = cv2.imread(file_name)
    demonstrate_file = file_name + ' plot_to_show.txt'
    demo_width = int(img.shape[1] * 0.0459)
    demo_height = int(img.shape[0] * 0.0235)
    img_re = cv2.resize(img, (demo_width, demo_height), interpolation=cv2.INTER_CUBIC)
    gray_img = grayscale(img_re)
    #N_str = '@%$&^!+=-. '
    N_str = '@%$&#*^!+  ' #New
    Li_index = gray_img/gray_img.max() * (len(N_str) - 1)
    Li_index = np.array(Li_index, dtype='int')
   
    #Show on Terminal and Text
    with open(demonstrate_file , 'w') as file:
        for j in range(demo_height):
                for i in range(demo_width):
                    num = Li_index[j, i]
                    print(N_str[num], end='')
                    file.write(N_str[num])
                print()
                file.write('\n')


def image_to_ascii(file_root, width_factor=1, height_factor=1):

    """
    從圖片轉成灰度圖片並利用明暗度轉換成ASCII 字符保存為 CSV 格式。

    參數:
    - file_root (str): 圖片 文件的路徑。
    - width_factor (float): 圖片寬度縮放比例。
    - height_factor (float): 圖片長度縮放比例。
    """

    image_to_asciitxt(file_root)

    # 讀取並調整圖片尺寸
    img = cv2.imread(file_root)
    re_width = int(img.shape[1] * width_factor)
    re_height = int(img.shape[0] * height_factor)
    img_re = cv2.resize(img, (re_width, re_height), interpolation=cv2.INTER_CUBIC)
    gray_img = grayscale(img_re)

    # 灰度值轉換為 ASCII 字符
    #N_str = '@%$&^!+=-. '
    N_str = '@%$&#*^!+  ' #New
    Li_index = gray_img/gray_img.max() * (len(N_str) - 1)
    Li_index = np.array(Li_index, dtype='int')
    output_file_path = 'output.csv'

    loguru.logger.info("Processing text output to CSV")
    with open(output_file_path, 'w', newline='') as file:
        writer = csv.writer(file)
        for j in range(re_height):
            row = [N_str[Li_index[j, i]] for i in range(re_width)]
            writer.writerow(row)

    loguru.logger.info("ASCII 字符畫已保存至 " + output_file_path)
    return output_file_path

        

def image_to_asciiColorful(file_root, width_factor=1, height_factor=1):

    """
    從圖片抽取顏色資訊以及轉成灰度圖片並利用明暗度轉換成ASCII 字符保存為 EXCEL 格式(含顏色與明暗度字符)。

    參數:
    - file_root (str): 圖片 文件的路徑。
    - width_factor (float): 圖片寬度縮放比例。
    - height_factor (float): 圖片長度縮放比例。
    """
    # 讀取並調整圖片尺寸
    img = cv2.imread(file_root)
    re_width = int(img.shape[1] * width_factor)
    re_height = int(img.shape[0] * height_factor)
    img_re = cv2.resize(img, (re_width, re_height), interpolation=cv2.INTER_CUBIC)
    gray_img = grayscale(img_re)
    img_re = cv2.cvtColor(img_re, cv2.COLOR_BGR2RGB)
    N_str = '@%$&#*^!+  ' #New

    # 灰度值轉換為 ASCII 字符
    Li_index = gray_img / gray_img.max() * (len(N_str) - 1)
    Li_index = np.array(Li_index, dtype='int')
    
    # 創建 Excel 工作簿
    workbook = Workbook()
    sheet = workbook.active
    output_file_path = 'outputColorful.xlsx'
    
    loguru.logger.info("Processing text output with color to Excel")

    for j in range(re_height):
        for i in range(re_width):
            ascii_char = N_str[Li_index[j, i]]
            # 提取該位置像素的 RGB 顏色
            r, g, b = img_re[j, i]
            # 將顏色轉換為十六進制格式
            hex_color = f"{r:02X}{g:02X}{b:02X}"

            # 在 Excel 中填入字符和背景顏色
            cell = sheet.cell(row=j+1, column=i+1, value=ascii_char)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            loguru.logger.debug(f'Exacute row, column:{j+1, i+1}')
            loguru.logger.info(f'process_to_EXCEL_FILE{(r, g, b), cell , hex_color, "char symbol:"+ascii_char}')
            time.sleep(0.02)

    # 保存 Excel 文件
    workbook.save(output_file_path)
    loguru.logger.info("ASCII 字符畫已保存至 " + output_file_path)
    return output_file_path










